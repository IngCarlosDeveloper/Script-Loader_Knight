import openpyxl
import time
import subprocess
import win32com.client

# Rutas de los archivos
xlsx_file = r"C:\\Users\\user\\Documents\\Trabajo\\Mercancia\\utilitarios\\datos.xlsx"  # Cambia a la ruta de tu archivo .xlsx
txt_file = r"C:\\a2Softway\\Empre001\\REPORTS\\Inventariofisico.txt"    # Cambia a la ruta de tu archivo .txt
xlsm_file = r"C:\\Users\\user\\Documents\\Trabajo\\Mercancia\\utilitarios\\LINK_KNIGHT_ULTRA.xlsm"         # Cambia al nombre deseado para tu archivo .xlsm

# Abrir el archivo .xlsm existente
workbook = openpyxl.load_workbook(xlsm_file, keep_vba=True)

# 1. Procesar el contenido del .xlsx y colocarlo en la Hoja 1
if "Hoja1" not in workbook.sheetnames:
    sheet1 = workbook.create_sheet(title="Hoja1")  # Crear la hoja si no existe
else:
    sheet1 = workbook["Hoja1"]  # Usar la hoja existente

# Limpiar la hoja antes de escribir (opcional, para evitar datos duplicados)
for row in sheet1.iter_rows():
    for cell in row:
        cell.value = None

# Leer el contenido del archivo .xlsx y escribir en la Hoja 1
xlsx_workbook = openpyxl.load_workbook(xlsx_file)
xlsx_sheet = xlsx_workbook.active

for row_idx, row in enumerate(xlsx_sheet.iter_rows(values_only=True), start=1):
    for col_idx, value in enumerate(row, start=1):
        cell = sheet1.cell(row=row_idx, column=col_idx, value=value)

        # Configurar la primera columna como texto
        if col_idx == 1:
            sheet1.cell(row=row_idx, column=col_idx).number_format = "@"

# 2. Procesar el contenido del .txt y colocarlo en la Hoja 2
if "Hoja2" not in workbook.sheetnames:
    sheet2 = workbook.create_sheet(title="Hoja2")  # Crear la hoja si no existe
else:
    sheet2 = workbook["Hoja2"]  # Usar la hoja existente

# Limpiar la hoja antes de escribir (opcional)
for row in sheet2.iter_rows():
    for cell in row:
        cell.value = None

# Leer el archivo .txt y escribir línea por línea en la Hoja 2
with open(txt_file, "r", encoding="latin-1") as txt:
    for line_idx, line in enumerate(txt, start=1):
        sheet2.cell(row=line_idx, column=1, value=line.strip())

# Guardar los cambios en el archivo .xlsm existente
workbook.save(xlsm_file)
print(f"Datos combinados y guardados en el archivo {xlsm_file}")

time.sleep(5)

#--------------------------------------------------EJECUTAR MACROS--------------------------------------------

# Ruta del archivo Excel
ruta_excel = r"C:\\Users\\user\\Documents\\Trabajo\\Mercancia\\utilitarios\\LINK_KNIGHT_ULTRA.xlsm"

# Crear una instancia de Excel
excel = win32com.client.Dispatch("Excel.Application")

# Abrir el archivo Excel
libro = excel.Workbooks.Open(ruta_excel)

try:
    # Ejecutar las macros en el orden deseado
    excel.Application.Run("LINK_KNIGHT_ULTRA.xlsm!MASTER")
    #excel.Application.Run("LINK_KNIGHT_ULTRA.xlsm!ThisWorkbook.CONCATENAR")
    #excel.Application.Run("LINK_KNIGHT_ULTRA.xlsm!ThisWorkbook.TextoEnColumnasConCodigos")
    #excel.Application.Run("LINK_KNIGHT_ULTRA.xlsm!ThisWorkbook.VerificarCodigosDeBarra")
    #excel.Application.Run("LINK_KNIGHT_ULTRA.xlsm!ThisWorkbook.ProcesarHoja3")
    #excel.Application.Run("LINK_KNIGHT_ULTRA.xlsm!ThisWorkbook.GuardarHoja3")


    print("Todas las macros se ejecutaron con éxito.")

except Exception as e:
    print(f"Error al ejecutar las macros: {e}")

finally:
    # Guardar y cerrar el archivo
    libro.Close(SaveChanges=True)
    excel.Application.Quit()

subprocess.run(["python", r"C:\\Users\\user\\Desktop\Loader_Knight\\cargador.py"])